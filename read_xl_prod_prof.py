import openpyxl as xl
import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
import numpy


def get_new_dates(start_date: datetime.datetime,end_date: datetime.datetime, period: str):
    if period not in ['monthly','quarterly','semi-annually','annually']:
        print("Incorrect period specification\nShould be one of 'monthly', 'quarterly', 'semi-annually', or 'annually'")
        return []
    if start_date > end_date:
        print("Start date cannot be later than End date\nExiting")
        return []
    dates=[]
    previous_date=start_date
    dates.append(previous_date)

    if period=='monthly':
        while True:
            next_date=previous_date+relativedelta(day=31)
            next_date+=relativedelta(days=1)
            next_date=next_date.replace(hour=0,minute=0,second=0,microsecond=0)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    elif period == 'quarterly':
        if start_date.month in [1,2,3]:
            next_date_month=4
            next_date_year=previous_date.year    
        elif start_date.month in [4,5,6]:
            next_date_month=7
            next_date_year=previous_date.year    
        elif start_date.month in [7,8,9]:
            next_date_month=10
            next_date_year=previous_date.year    
        else:
            next_date_month=1
            next_date_year=previous_date.year+1
        next_date=datetime.datetime(next_date_year,next_date_month,1)
        dates.append(next_date)
        previous_date=next_date
        while True:
            next_date=previous_date+relativedelta(months=3)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    elif period=='semi-annually':
        if start_date.month in [1,2,3,4,5,6]:
            next_date_month=7
            next_date_year=previous_date.year   
        else:
            next_date_month=1
            next_date_year=previous_date.year+1
        next_date=datetime.datetime(next_date_year,next_date_month,1)
        dates.append(next_date)
        previous_date=next_date
        while True:
            next_date=previous_date+relativedelta(months=6)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    elif period=='annually':
        while True:
            next_date=datetime.datetime(previous_date.year+1,1,1,0,0,0)
            if next_date == end_date:
                dates.append(next_date)
                break
            elif next_date > end_date:
                dates.append(end_date)
                break
            else:
                dates.append(next_date)
                previous_date=next_date
    return dates      


    if x<=x_arr[0]:
        return y_arr[0]
    if x>=x_arr[-1]:
        return y_arr[-1]

    for i in range(1,len(x_arr)-1):
        if x<=x_arr[i] and x>x_arr[i-1]:
            return y_arr[i-1]+((y_arr[i]-y_arr[i-1])/(x_arr[i]-x_arr[i-1]))*(x-x_arr[i-1])

def get_required_columns(input_df: pd.DataFrame):
  # df=pd.read_excel(workbook,sheet_name=worksheet,header=[0,1],index_col=0,parse_dates=True,engine='openpyxl')
  columns_to_delete=[]
  for tup in input_df.columns.values:
    if "CUM" not in tup[1]:
      columns_to_delete.append(tup)
  for col in columns_to_delete:
    input_df.pop(col)
  return input_df
    
def dates_bulk_shifted(input_dataframe,days_to_shift):

  required_columns=input_dataframe.columns.values
  data_values=input_dataframe.values
  data_arrays=[[data_values[i,j] for i in range(len(data_values))] for j in range(len(required_columns))]
  dates_list=input_dataframe.index.values

  days_from_start=[(i-dates_list[0])/numpy.timedelta64(1,'D') for i in dates_list]
  new_dates=[j+numpy.timedelta64(days_to_shift,'D') for j in dates_list]
  new_days_from_start=[(i-new_dates[0])/numpy.timedelta64(1,'D') for i in new_dates]

  sers = [pd.Series(numpy.interp(new_days_from_start,days_from_start,data_arrays[i])) for i in range(len(required_columns))]
  sers_with_rates=get_rates_from_cum(sers,required_columns,new_days_from_start)

  output_df=pd.DataFrame(pd.concat(sers,axis=1))
  output_df.index=new_dates
  output_df.columns=pd.MultiIndex.from_tuples(required_columns)
  # print(output_df.head)

  return output_df

def get_rates_from_cum(series,columns,days_from_start):
  new_columns=[]
  for i in range(len(columns)):
    new_columns.append(columns[i])
    rate_type="AVG._"+columns[i][1].replace('CUM','')+"_RATE"
    new_columns.append((columns[i][0],rate_type))
  
  new_series=[]
  for ser in series:
    new_series.append(ser)
    rate_series=[]
    for i in range(len(ser)-1):
      rate_series[i]=(ser[i+1]-ser[i])/(days_from_start[i+1]-days_from_start[i])
    rate_series.append(0)
    new_series.append(rate_series)
  
  return new_columns,new_series






def get_shift_duration(workbook):
  wbk=xl.load_workbook(workbook)
  worksheets=wbk.sheetnames
  if 'Bulk_Shift_Input' in worksheets:
    first_row=next(wbk['Bulk_Shift_Input'].rows)
    wbk.close()
    del wbk
    return(first_row[1].value)
  else:
    wbk.close()
    del wbk
    return None


print("Hello World")
shift_duration=get_shift_duration(r"C:\Users\rdandekar\Desktop\Prod_Prof.xlsx")
if shift_duration is None:
  exit()
entity_types=['USERDF','SOURCE','SEP','TANK','JOINT','WELL','INLGEN']
df_dict={}
wbk=xl.load_workbook(r"C:\Users\rdandekar\Desktop\Prod_Prof.xlsx")
for entity_type in entity_types:
  df=pd.DataFrame()
  df=pd.read_excel(wbk,sheet_name=entity_type,header=[0,1],index_col=0,parse_dates=True,engine='openpyxl')
  df=get_required_columns(df)
  
  if not df.empty:
    df=dates_bulk_shifted(df,shift_duration)
    df_dict[entity_type]=df

entity_types_shifted=[]
entity_types_shifted_df=[]
for key in df_dict.keys():
  entity_types_shifted.append(key)
  entity_types_shifted_df.append(df_dict[key])
entity_types_shifted=[ type+'_SHIFTED' for type in entity_types]



writer = pd.ExcelWriter(r"C:\Users\rdandekar\Desktop\Prod_Prof_shifted.xlsx", engine='openpyxl')
for entity, entity_df in zip(entity_types_shifted,entity_types_shifted_df):
  entity_df.to_excel(writer,sheet_name=entity)
writer.save()
writer.close()
exit()


# try:
#   workbook=xl.load_workbook(r"C:\Users\rdandekar\Desktop\Prod_Prof.xlsx")
#   print(type(workbook))
# except Exception:
#   print('File not found')
#   exit()
# print(f'Workbook opened \n')


# read bulk shift days from excel_input_sheet
# first_row=next(workbook['Bulk_Shift_Input'].rows)
# days_to_shift=first_row[1].value
# worksheets=workbook.sheetnames
# workbook.close()
# writer=pd.ExcelWriter(r"C:\Users\rdandekar\Desktop\Prod_Prof_shifted.xlsx",engine='openpyxl')
# writer.book=workbook
#for testing
# entity_types=['USERDF','SOURCE','SEP','TANK','JOINT','WELL','INLGEN']
# entity_dates=[]


# for entity_type in entity_types:
#   df=pd.DataFrame()
#   df=dates_bulk_shifted(r"C:\Users\rdandekar\Desktop\Prod_Prof.xlsx",entity_type,days_to_shift)
#   if df is not None:
#     df.to_excel(writer,sheet_name=entity_type+"_shifted")
# writer.save()
# writer.close()

# workbook.save()
# workbook.close()
# print("I am here")

